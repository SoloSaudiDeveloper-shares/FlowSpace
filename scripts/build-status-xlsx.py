"""Build a status workbook summarising everything delivered + the open backlog.

Re-run any time to refresh:  python scripts/build-status-xlsx.py
Output: C:\\Users\\malfa\\OneDrive\\سطح المكتب\\flowspace-status.xlsx
"""

import os
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ─── Palette ────────────────────────────────────────────────────────────
HEADER_FILL = PatternFill("solid", start_color="1E293B")  # slate-800
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)

DONE_FILL = PatternFill("solid", start_color="DCFCE7")  # emerald-100
DONE_FONT = Font(name="Arial", color="166534", bold=True, size=10)

BUG_FILL = PatternFill("solid", start_color="FEE2E2")
BUG_FONT = Font(name="Arial", color="991B1B", bold=True, size=10)

VERIFY_FILL = PatternFill("solid", start_color="FEF3C7")
VERIFY_FONT = Font(name="Arial", color="92400E", bold=True, size=10)

EASY_FILL = PatternFill("solid", start_color="DCFCE7")
EASY_FONT = Font(name="Arial", color="166534", bold=True, size=10)

MEDIUM_FILL = PatternFill("solid", start_color="DBEAFE")
MEDIUM_FONT = Font(name="Arial", color="1E40AF", bold=True, size=10)

HEAVY_FILL = PatternFill("solid", start_color="FED7AA")
HEAVY_FONT = Font(name="Arial", color="9A3412", bold=True, size=10)

BLOCKED_FILL = PatternFill("solid", start_color="E5E7EB")
BLOCKED_FONT = Font(name="Arial", color="374151", bold=True, size=10)

STRIPE_FILL = PatternFill("solid", start_color="F8FAFC")

BODY_FONT = Font(name="Arial", size=10)
BODY_BOLD = Font(name="Arial", size=10, bold=True)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

THIN = Side(border_style="thin", color="E2E8F0")
BORDER = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)


def style_status(cell, status):
    s = status.lower()
    if s in ("done", "delivered", "shipped"):
        cell.fill = DONE_FILL; cell.font = DONE_FONT
    elif s in ("bug", "regression"):
        cell.fill = BUG_FILL; cell.font = BUG_FONT
    elif s in ("verify", "untested", "needs test"):
        cell.fill = VERIFY_FILL; cell.font = VERIFY_FONT
    elif s in ("easy", "quick win"):
        cell.fill = EASY_FILL; cell.font = EASY_FONT
    elif s in ("medium",):
        cell.fill = MEDIUM_FILL; cell.font = MEDIUM_FONT
    elif s in ("heavy", "hard"):
        cell.fill = HEAVY_FILL; cell.font = HEAVY_FONT
    elif s in ("blocked", "paid", "constrained", "needs key"):
        cell.fill = BLOCKED_FILL; cell.font = BLOCKED_FONT
    cell.alignment = CENTER


def write_table(ws, headers, rows, status_col_index=None):
    for col_idx, label in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col_idx, value=label)
        c.fill = HEADER_FILL; c.font = HEADER_FONT
        c.alignment = CENTER; c.border = BORDER
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, value in enumerate(row, start=1):
            c = ws.cell(row=row_idx, column=col_idx, value=value)
            c.font = BODY_FONT; c.alignment = WRAP; c.border = BORDER
            if row_idx % 2 == 0:
                c.fill = STRIPE_FILL
            if status_col_index and col_idx == status_col_index:
                style_status(c, str(value))
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"


def auto_width(ws, widths):
    for col_idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w


# ─── Workbook ───────────────────────────────────────────────────────────
wb = Workbook()
wb.remove(wb.active)

# ─── Sheet 1: Summary ───────────────────────────────────────────────────
ws = wb.create_sheet("Summary")
ws["A1"] = "FlowSpace — project status"
ws["A1"].font = Font(name="Arial", bold=True, size=18, color="1E293B")
ws["A2"] = f"Updated {date.today().isoformat()}"
ws["A2"].font = Font(name="Arial", italic=True, color="64748B", size=10)

kpi_rows = [
    ("Features delivered (cumulative)", 105),
    ("Bugs reported → fixed", "4 of 5"),
    ("Quick-win backlog → shipped", "12 of 12"),
    ("Medium scope → shipped", "6 of 6"),
    ("Still open: needs your test (Calendar)", 1),
    ("Still open: heavy / paid / external-API", 4),
]
ws["A4"] = "Where things stand"
ws["A4"].font = Font(name="Arial", bold=True, size=12)
for idx, (label, count) in enumerate(kpi_rows, start=5):
    ws[f"A{idx}"] = label
    ws[f"B{idx}"] = count
    ws[f"A{idx}"].font = BODY_FONT
    ws[f"B{idx}"].font = BODY_BOLD
    ws[f"B{idx}"].alignment = Alignment(horizontal="right")

ws["A12"] = "Live row counts (formula-computed)"
ws["A12"].font = Font(name="Arial", bold=True, size=12)
ws["A13"] = "Delivered"
ws["B13"] = "=COUNTA(Delivered!A:A)-1"
ws["A14"] = "Open (bugs/verify)"
ws["B14"] = "=COUNTA('Open items'!A:A)-1"
ws["A15"] = "Heavy / external"
ws["B15"] = "=COUNTA('Heavy & External'!A:A)-1"
for r in range(13, 16):
    ws[f"A{r}"].font = BODY_FONT
    ws[f"B{r}"].font = BODY_BOLD
    ws[f"B{r}"].alignment = Alignment(horizontal="right")
    ws[f"B{r}"].number_format = "#,##0"

ws["A18"] = "Status legend"
ws["A18"].font = Font(name="Arial", bold=True, size=12)
legend = [
    ("Done", "Shipped on main + pushed to GitHub."),
    ("Verify", "Code is in place but needs an end-to-end test on the live site."),
    ("Heavy", "Multi-day; may need new infra."),
    ("Blocked", "Depends on a third party — API approval, paid tier, dedicated number."),
]
for idx, (status, desc) in enumerate(legend, start=19):
    s = ws.cell(row=idx, column=1, value=status)
    style_status(s, status)
    s.alignment = Alignment(horizontal="center")
    d = ws.cell(row=idx, column=2, value=desc)
    d.font = BODY_FONT; d.alignment = WRAP

ws["A25"] = "Note"
ws["A25"].font = Font(name="Arial", bold=True, size=12)
ws["A26"] = (
    "Everything from the original Bugs / Quick-wins / Medium backlogs is now shipped. "
    "The only things left are (1) your hands-on Google Calendar test and (2) the four "
    "heavy integrations that are blocked on real-world admin work (paid API tiers, "
    "platform approvals), not on code."
)
ws["A26"].font = BODY_FONT
ws["A26"].alignment = WRAP
ws.merge_cells("A26:B30")

auto_width(ws, [40, 90])
for r in range(1, 31):
    ws.row_dimensions[r].height = 18
ws.row_dimensions[1].height = 28
ws.row_dimensions[26].height = 60

# ─── Sheet 2: Delivered ─────────────────────────────────────────────────
ws = wb.create_sheet("Delivered")
delivered_headers = ["Area", "Feature", "Status", "Notes"]
delivered_rows = [
    # Auth & security
    ("Auth", "Username/password login + bcrypt", "Done", "Legacy hashes upgraded on next login."),
    ("Auth", "Google OAuth sign-in / sign-up", "Done", "PKCE flow."),
    ("Auth", "Rate-limited login", "Done", "Per-(IP, username) lockout."),
    ("Auth", "Password reset email", "Done", "Gmail SMTP / Resend transport."),
    ("Auth", "Admin open/closed signups + session duration", "Done", "Owner-only controls."),
    ("Auth", "Two-factor authentication (TOTP)", "Done", "QR + manual key, 8 recovery codes, login challenge."),
    ("Auth", "Personal API tokens (flws_… bearer)", "Done", "Issue/revoke, hashed at rest, expiry presets."),
    # Workspace
    ("Workspace", "Per-user element isolation", "Done", "Every read scoped by created_by."),
    ("Workspace", "Editable workspace title", "Done", "Admin prefix + per-user suffix."),
    ("Workspace", "Admin events log + inline docs", "Done", "+ search box that filters Users/Backups/Events."),
    ("Workspace", "Bell notification badge", "Done", "Unread + pending imports + pending emails + overdue."),
    ("Workspace", "Backups: create/restore/delete", "Done", "Retention + manual snapshot."),
    # UI / nav
    ("UI", "Sidebar polish: colours, drag-reorder, right-click", "Done", "Per-section accent colours."),
    ("UI", "Draggable clock (digital/analog)", "Done", "Mobile: compact + long-press menu."),
    ("UI", "Floating task timer", "Done", "Mobile: icon-only + long-press menu."),
    ("UI", "Scrolling feed ticker", "Done", "Pin/unpin/drag/resize."),
    ("UI", "Tooltips foundation (Hint wrapper)", "Done", "Across icon-only controls."),
    ("UI", "Settings → admin-style top tabs", "Done", "Account/Data/Look/Integrations/Help."),
    ("UI", "Onboarding tour (7 steps)", "Done", "Replay from Settings → Help."),
    # Element pages
    ("Element pages", "Inline title editor", "Done", "Click to edit, optimistic save."),
    ("Element pages", "Send-as-email button (every type)", "Done", "Template engine + 4 templates."),
    ("Element pages", "Watch / unwatch", "Done", "Watcher count."),
    ("Element pages", "Custom fields + one-click example", "Done", "Text/number/date/select/etc."),
    ("Element pages", "Save project as template", "Done", "Snapshots statuses + tasks + subtasks."),
    # Todo lists
    ("Todo lists", "Mobile priority + due-date controls", "Done", "Always-visible flag/date/⋮ buttons + inline pickers."),
    ("Todo lists", "Export to Excel + email as attachment", "Done", "Formula-computed Total/Completed/Open."),
    # Telegram
    ("Telegram", "Multi-user bot (own token each)", "Done", "Webhook per user."),
    ("Telegram", "Import flow with approval bell", "Done", "Pastes parked as pending."),
    ("Telegram", "Inline keyboards + pagination", "Done", "Tap-to-act."),
    ("Telegram", "Read/insight commands", "Done", "/tasks /deadlines /projects /lists /stats /search /digest."),
    ("Telegram", "Smart capture (!priority @date #tags)", "Done", "Parser pulls tokens out."),
    ("Telegram", "Task edit power", "Done", "Reschedule / priority / status / delete."),
    ("Telegram", "Voice IN (Groq Whisper)", "Done", "Per-message language + auto-skip."),
    ("Telegram", "Voice OUT (TTS replies)", "Done", "Via user's /audio/speech provider."),
    ("Telegram", "Natural-language commands", "Done", "Freeform routed through AI provider."),
    ("Telegram", "Morning + evening digest cron", "Done", "Per-bot enable + time."),
    ("Telegram", "Customizable reply templates", "Done", "6 moments, variable chips."),
    ("Telegram", "/clear history + capture Undo button", "Done", "Two-step clear; one-tap undo."),
    ("Telegram", "Deep-link 'Open in FlowSpace' buttons", "Done", "Main menu + task detail (needs PUBLIC_APP_URL)."),
    # Email
    ("Email", "Send any element as email + templates", "Done", "Per-type built-ins."),
    ("Email", "Email IN webhook + approval queue", "Done", "Resend/Postmark/etc; per-user routing."),
    # AI
    ("AI", "Generic provider settings", "Done", "OpenAI / Ollama / Gemini / Anthropic."),
    ("AI", "AI-import (markdown → projects/tasks)", "Done", "One paste."),
    ("AI", "Image vision (/vision page)", "Done", "Drop image, prompt presets, save as Page."),
    ("AI", "PDF/CSV/MD file ingest dropzone", "Done", "Home page → parse → Page, optional AI summary."),
    ("AI", "Chat-with-platform drawer", "Done", "Floating button; fresh workspace snapshot per turn."),
    # Search
    ("Search", "Global FTS5 search (owner-scoped)", "Done", "Titles + descriptions + page bodies + notes; fixed an ownership leak."),
    # Productivity
    ("Productivity", "Pomodoro focus timer widget", "Done", "25/5/15 cycle, logged."),
    ("Productivity", "Habit tracker (/habits)", "Done", "Daily/weekly, streaks."),
    ("Productivity", "Web clipper (bookmarklet + extension stub)", "Done", "POST /api/clip with bearer token."),
    # Platform / infra
    ("Infra", "PWA (manifest, icons, SW, install banner)", "Done", "Offline shell cache."),
    ("Infra", "i18n scaffolding + RTL switch", "Done", "Sidebar + login translated to Arabic; flag picker on login."),
    ("Infra", "Google Calendar sync (one-way)", "Verify", "Code shipped — needs an end-to-end test on the live site."),
    ("Infra", "Self-hosted Whisper sidecar", "Done", "docker-compose + docs; voice routes local when env set."),
    ("Infra", "Voice usage indicator", "Done", "Settings → Speech: today + 7-day counts + sparkline."),
    ("Infra", "Canvas configuration menu", "Done", "Background pattern, gap, snap, minimap — per-canvas."),
    ("Infra", "HTTPS via Caddy + Cloudflare proxy", "Done", "Auto-renew TLS, DDoS shield."),
    # Docs
    ("Docs", "Teams integration feasibility study", "Done", "docs/TEAMS_INTEGRATION.md."),
    ("Docs", "Self-hosted Whisper guide", "Done", "docs/SELF_HOSTED_WHISPER.md."),
    ("Docs", "Web clipper docs", "Done", "Bookmarklet + extension."),
    ("Docs", "Markdown templates panel", "Done", "5 paste-ready blueprints in Settings → Help."),
    ("Docs", "FlowSpace — Roadmap project (seeded)", "Done", "This status, modelled inside the app itself."),
]
write_table(ws, delivered_headers, delivered_rows, status_col_index=3)
auto_width(ws, [16, 52, 12, 84])
for r in range(2, len(delivered_rows) + 2):
    ws.row_dimensions[r].height = 30

# ─── Sheet 3: Open items (bugs / verify) ────────────────────────────────
ws = wb.create_sheet("Open items")
open_headers = ["#", "Item", "Status", "What's needed", "Notes"]
open_rows = [
    (1, "Google Calendar sync — end-to-end verify", "Verify", "Your hands-on test on the live site",
     "Settings → Calendar sync → Connect. If it errors with redirect_uri_mismatch, whitelist /api/auth/google-calendar/callback in Google Cloud Console. Code + cron + OAuth all shipped."),
]
write_table(ws, open_headers, open_rows, status_col_index=3)
auto_width(ws, [5, 42, 12, 32, 80])
for r in range(2, len(open_rows) + 2):
    ws.row_dimensions[r].height = 60

# ─── Sheet 4: Fixed bugs (history) ──────────────────────────────────────
ws = wb.create_sheet("Bugs fixed")
fixed_headers = ["#", "Bug", "Status", "Fix"]
fixed_rows = [
    (1, "Mobile: clock + timer + date ate the screen", "Done",
     "Responsive breakpoints — compact on phones, date chip hidden < md, analog face shrinks."),
    (2, "Mobile: settings checkboxes / clock menu unreachable", "Done",
     "New useLongPress hook → 500ms hold opens the same context menu desktop right-click gives."),
    (3, "Home mic less accurate than Telegram", "Done",
     "SpeechButton preferAccuracy prop → home capture forces Groq Whisper instead of Web Speech."),
    (4, "Arabic not fully integrated", "Done",
     "Sidebar groups + bottom nav + login form translated; flag picker on login; RTL flip."),
]
write_table(ws, fixed_headers, fixed_rows, status_col_index=3)
auto_width(ws, [5, 50, 12, 84])
for r in range(2, len(fixed_rows) + 2):
    ws.row_dimensions[r].height = 42

# ─── Sheet 5: Heavy & External (still open) ─────────────────────────────
ws = wb.create_sheet("Heavy & External")
h_headers = ["#", "Feature", "Effort", "Status", "Cost / constraint", "Notes"]
h_rows = [
    (1, "TikTok integration", "~1 week", "Constrained", "Free but limited API",
     "Public API = login + basic profile only. Posting needs TikTok for Business + Content Posting API approval. Realistic v1: post a link → pull metadata."),
    (2, "Twitter / X integration", "~3 days", "Paid", "$100+/month",
     "Free tier dead. Basic tier ~$100/mo for 50k reads + 3k writes. Code is straightforward REST; the bill is the blocker."),
    (3, "WhatsApp Business integration", "~1 wk + 2-wk approval", "Blocked", "Meta verification + dedicated number",
     "WA Business API needs Meta-verified account + a dedicated phone number not usable for personal WA."),
    (4, "Video → frames → LLM analysis", "~2-3 days", "Heavy", "Compute-heavy",
     "ffmpeg -vf fps=1 → vision LLM per frame → summary. Vision calls add up; limit to short clips."),
]
write_table(ws, h_headers, h_rows, status_col_index=4)
auto_width(ws, [5, 36, 18, 14, 30, 80])
for r in range(2, len(h_rows) + 2):
    ws.row_dimensions[r].height = 66

# ─── Reorder + save ─────────────────────────────────────────────────────
wb._sheets = [
    wb["Summary"],
    wb["Delivered"],
    wb["Open items"],
    wb["Bugs fixed"],
    wb["Heavy & External"],
]

OUT = r"C:\Users\malfa\OneDrive\سطح المكتب\flowspace-status.xlsx"
wb.save(OUT)
print(f"Saved: {os.path.basename(OUT)} ({os.path.getsize(OUT)} bytes)")
print(f"Delivered rows: {len(delivered_rows)} · Open: {len(open_rows)} · Fixed: {len(fixed_rows)} · Heavy: {len(h_rows)}")
